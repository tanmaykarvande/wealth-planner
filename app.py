import streamlit as st
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import os
import tempfile
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, Image as RLImage, HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors


# ══════════════════════════════════════════════════════════════════════════════
# INDIAN NUMBER FORMATTING UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def indian_fmt(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "Rs.0"
    x = float(x)
    neg = x < 0
    x = abs(x)
    if x >= 1e7:
        s = f"Rs.{x/1e7:.2f}Cr"
    elif x >= 1e6:
        s = f"Rs.{x/1e5:.2f}L"
    else:
        s = f"Rs.{x:,.0f}"
    return f"-{s}" if neg else s

def indian_fmt_short(x):
    if x is None: return "0"
    x = float(x)
    if x == 0: return "0"
    neg = x < 0
    x = abs(x)
    if x >= 1e7:
        s = f"{x/1e7:.1f}Cr"
    elif x >= 1e6:
        s = f"{x/1e5:.1f}L"
    else:
        s = f"{x:,.0f}"
    return f"-{s}" if neg else s

# ══════════════════════════════════════════════════════════════════════════════
# 1. GRANULAR WEALTH ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class GranularWealthEngine:
    def __init__(self, inputs: dict, step_up_schedule: dict):
        self.client_name       = inputs.get("client_name", "Valued Client")
        self.current_age       = inputs.get("current_age", 35)
        self.annual_premium    = inputs.get("annual_premium", 150_000)
        self.payout_pct        = inputs.get("payout_pct", 0.40)
        self.life_cover_mult   = inputs.get("life_cover_multiple", 7)
        self.ppt               = inputs.get("ppt", 12)
        self.policy_term       = inputs.get("policy_term", 40)
        self.expected_return   = inputs.get("expected_return", 0.18)
        self.monthly_swp       = inputs.get("monthly_swp", 100_000)
        self.swp_start_age     = inputs.get("swp_start_age", 60)
        self.step_up_schedule  = step_up_schedule

        self.life_cover        = self.annual_premium * self.life_cover_mult
        self.annual_payout     = self.annual_premium * self.payout_pct
        self.base_monthly_sip  = round(self.annual_payout / 12, 2)
        self.monthly_rate      = self.expected_return / 12

    def run_projection(self) -> pd.DataFrame:
        records = []
        corpus  = 0.0
        cumulative_step_up = 0.0

        for year in range(1, self.policy_term + 1):
            age = self.current_age + (year - 1)

            premium_paid      = self.annual_premium if year <= self.ppt else 0.0
            insurance_payout  = self.annual_payout

            year_increment     = self.step_up_schedule.get(year, 0.0)
            cumulative_step_up += year_increment
            total_monthly_sip  = self.base_monthly_sip + cumulative_step_up
            annual_sip_contrib = total_monthly_sip * 12

            monthly_swp = self.monthly_swp if age >= self.swp_start_age else 0.0

            for _ in range(12):
                corpus = max(corpus, 0.0)
                corpus += total_monthly_sip
                corpus *= (1 + self.monthly_rate)
                if corpus >= monthly_swp:
                    corpus -= monthly_swp
                else:
                    corpus = 0.0

            net_corpus       = round(corpus, 2)
            annual_swp_drawn = (self.monthly_swp * 12) if age >= self.swp_start_age else 0.0
            status           = "Corpus Exhausted" if net_corpus <= 0 else "Sustainable"

            records.append({
                "Policy Year":                   year,
                "Age":                           age,
                "Premium Paid (Rs.)":            premium_paid,
                "Insurance Payout (Rs.)":        insurance_payout,
                "Base Monthly SIP (Rs.)":        self.base_monthly_sip,
                "New Step-Up Added (Rs.)":       year_increment,
                "Total Monthly SIP (Rs.)":       total_monthly_sip,
                "Annual SIP Contribution (Rs.)": annual_sip_contrib,
                "Net End-of-Year Corpus (Rs.)":  net_corpus,
                "Annual SWP Withdrawal (Rs.)":   annual_swp_drawn,
                "Sustainability Flag":           status,
            })

        return pd.DataFrame(records)

    def generate_summary(self, df: pd.DataFrame) -> dict:
        ret_row = df[df["Age"] == self.swp_start_age]
        corpus_at_ret = (
            ret_row["Net End-of-Year Corpus (Rs.)"].values[0]
            if not ret_row.empty else 0.0
        )
        final_corpus = df["Net End-of-Year Corpus (Rs.)"].iloc[-1]

        n_months = (self.policy_term - (self.swp_start_age - self.current_age)) * 12
        r = self.monthly_rate
        required_corpus = (
            self.monthly_swp * ((1 - (1 + r) ** -n_months) / r)
            if n_months > 0 else 0.0
        )

        exhausted = df[df["Sustainability Flag"] == "Corpus Exhausted"]
        if not exhausted.empty:
            sustain_status = "Corpus Exhausted"
            survives_yr    = exhausted["Policy Year"].values[0]
            survives_age   = exhausted["Age"].values[0]
        else:
            sustain_status = "Sustainable"
            survives_yr    = self.policy_term
            survives_age   = self.current_age + self.policy_term

        if sustain_status == "Sustainable" and corpus_at_ret > 0 and n_months > 0:
            max_swp = corpus_at_ret * r / (1 - (1 + r) ** -n_months)
        else:
            max_swp = 0.0

        safe_note = (
            f"Based on the projected corpus of {indian_fmt(corpus_at_ret)} at age {self.swp_start_age}, "
            f"the strategy can safely sustain a monthly withdrawal of {indian_fmt(max_swp)} "
            f"across the remaining {n_months // 12}-year horizon."
        ) if sustain_status == "Sustainable" else (
            f"⚠️ The current corpus trajectory is insufficient. "
            f"Corpus exhaustion is projected at Policy Year {survives_yr} (Age {survives_age})."
        )

        return {
            "Life Cover Amount":        self.life_cover,
            "Monthly Insurance Payout": self.base_monthly_sip,
            "Corpus at SWP Start":      corpus_at_ret,
            "Target Required Corpus":   round(required_corpus, 2),
            "Surplus / Shortfall":      round(corpus_at_ret - required_corpus, 2),
            "Final Year Corpus":        final_corpus,
            "Sustainability Status":    sustain_status,
            "Years Corpus Survives":    survives_yr,
            "Age Corpus Exhausted":     survives_age,
            "Max Safe Monthly SWP":     round(max_swp, 2),
            "Client Presentation Note": safe_note,
        }


# ══════════════════════════════════════════════════════════════════════════════
# 2. CHART BUILDER  — Indian number formatting on axes and annotations
# ══════════════════════════════════════════════════════════════════════════════

def build_chart(df: pd.DataFrame, swp_start_age: int) -> plt.Figure:
    fig, ax = plt.subplots(figsize=(9, 4), facecolor="#0F1923")
    ax.set_facecolor("#0F1923")

    ages   = df["Age"]
    corpus = df["Net End-of-Year Corpus (Rs.)"]

    swp_mask = ages >= swp_start_age
    ax.fill_between(ages, corpus, where=~swp_mask, alpha=0.08, color="#4A90E2")
    ax.fill_between(ages, corpus, where=swp_mask,  alpha=0.12, color="#E05C2A")

    ax.plot(ages[~swp_mask], corpus[~swp_mask], color="#4A90E2", lw=2.2, label="Accumulation Phase")
    ax.plot(ages[swp_mask],  corpus[swp_mask],  color="#E05C2A", lw=2.2, label="SWP Phase")

    # Annotate every 5 years — Indian format
    for _, row in df[df["Age"] % 5 == 0].iterrows():
        val = row["Net End-of-Year Corpus (Rs.)"]
        ax.annotate(
            indian_fmt_short(val),
            (row["Age"], val),
            fontsize=7.5, xytext=(0, 8), textcoords="offset points",
            ha="center", color="#C5D8EE", fontfamily="monospace",
        )

    ax.axvline(swp_start_age, color="#E05C2A", lw=1, linestyle="--", alpha=0.6)

    # Y-axis ticks — Indian format
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: indian_fmt_short(x))
    )

    ax.set_xlabel("Client Age", fontsize=9, color="#8BA3BF")
    ax.set_ylabel("Portfolio Value", fontsize=9, color="#8BA3BF")
    ax.tick_params(labelsize=8, colors="#8BA3BF")
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines["left"].set_color("#2A3F55")
    ax.spines["bottom"].set_color("#2A3F55")
    ax.legend(fontsize=8, framealpha=0, labelcolor="#C5D8EE")
    ax.grid(axis="y", alpha=0.2, linestyle="--", color="#4A90E2")
    fig.tight_layout()
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# 3. EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def _thin(color="DDDDDD"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def build_excel(df, summary, inputs, step_up_schedule, chart_path) -> bytes:
    wb = openpyxl.Workbook()

    ws_sum  = wb.active;          ws_sum.title  = "Executive Summary"
    ws_proj = wb.create_sheet("40-Year Projection")
    ws_gr   = wb.create_sheet("Charts")

    NAVY  = "1B365D"
    LIGHT = "F0F4F8"
    WHITE = "FFFFFF"
    G_FILL = "C6EFCE"; G_TEXT = "006100"
    R_FILL = "FFC7CE"; R_TEXT = "9C0006"

    F_TITLE = Font(name="Calibri", size=15, bold=True,  color=NAVY)
    F_SUB   = Font(name="Calibri", size=10, italic=True, color="555555")
    F_HDR   = Font(name="Calibri", size=10, bold=True,  color=WHITE)
    F_BOLD  = Font(name="Calibri", size=10, bold=True)
    F_REG   = Font(name="Calibri", size=10)
    A_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    A_LEFT  = Alignment(horizontal="left",   vertical="center")
    A_RIGHT = Alignment(horizontal="right",  vertical="center")

    # ── Sheet 1: Executive Summary ────────────────────────────────────────────
    ws_sum.column_dimensions["B"].width = 36
    ws_sum.column_dimensions["C"].width = 28

    ws_sum.merge_cells("B1:C1")
    t = ws_sum["B1"]
    t.value = "WEALTH PLANNER EXECUTIVE DASHBOARD"
    t.font  = F_TITLE; t.alignment = A_LEFT

    ws_sum["B2"].value = f"Strategic Advisory Matrix — {inputs['client_name']}"
    ws_sum["B2"].font  = F_SUB

    for col, label in [("B", "Parameter"), ("C", "Value")]:
        c = ws_sum[f"{col}4"]
        c.value = label; c.font = F_HDR
        c.fill  = _fill(NAVY); c.alignment = A_CTR; c.border = _thin()

    is_bad = summary["Sustainability Status"] != "Sustainable"

    kpis = [
        ("Client Name",                   inputs["client_name"],                    "text"),
        ("Current Age",                   inputs["current_age"],                    "int"),
        ("Policy Premium Term (years)",   inputs["ppt"],                            "int"),
        ("Expected Annual Return",        inputs["expected_return"],                "pct"),
        ("SWP Start Age",                 inputs["swp_start_age"],                  "int"),
        ("Initial Monthly SIP",           indian_fmt(summary["Monthly Insurance Payout"]), "text"),
        ("Total Life Cover",              indian_fmt(summary["Life Cover Amount"]),  "text"),
        (f"Corpus at Age {inputs['swp_start_age']}", indian_fmt(summary["Corpus at SWP Start"]), "text"),
        ("Target Corpus for SWP",         indian_fmt(summary["Target Required Corpus"]), "text"),
        ("Surplus / Shortfall",           indian_fmt(summary["Surplus / Shortfall"]), "text"),
        ("Max Safe Monthly Withdrawal",   indian_fmt(summary["Max Safe Monthly SWP"]), "text"),
        ("Terminal Portfolio (Yr End)",   indian_fmt(summary["Final Year Corpus"]),  "text"),
        ("Strategy Sustainability",       summary["Sustainability Status"].upper(),  "status"),
    ]

    row = 5
    for label, val, vtype in kpis:
        ck = ws_sum.cell(row=row, column=2, value=label)
        cv = ws_sum.cell(row=row, column=3, value=val)
        ck.font = F_REG; ck.border = _thin(); cv.border = _thin()
        if row % 2 == 0:
            ck.fill = _fill(LIGHT); cv.fill = _fill(LIGHT)
        if vtype == "pct":
            cv.number_format = '0.00%'; cv.alignment = A_RIGHT
        elif vtype == "int":
            cv.number_format = '#,##0';  cv.alignment = A_CTR
        elif vtype == "status":
            fill_c = R_FILL if is_bad else G_FILL
            text_c = R_TEXT if is_bad else G_TEXT
            cv.fill = _fill(fill_c)
            cv.font = Font(name="Calibri", size=10, bold=True, color=text_c)
            cv.alignment = A_CTR
        else:
            cv.alignment = A_RIGHT
        row += 1

    row += 1
    ws_sum.cell(row=row, column=2, value="Step-Up Schedule Applied").font = F_BOLD
    row += 1
    if step_up_schedule:
        for yr, amt in sorted(step_up_schedule.items()):
            ws_sum.cell(row=row, column=2, value=f"  Year {yr}").font = F_REG
            c = ws_sum.cell(row=row, column=3, value=indian_fmt(amt))
            c.font = F_REG; c.alignment = A_RIGHT
            row += 1
    else:
        ws_sum.cell(row=row, column=2, value="  No step-up configured").font = F_SUB
        row += 1

    row += 1
    ws_sum.merge_cells(start_row=row, start_column=2, end_row=row+3, end_column=3)
    nc = ws_sum.cell(row=row, column=2, value=f"💡 {summary['Client Presentation Note']}")
    nc.font = Font(name="Calibri", size=9, italic=True, color=NAVY)
    nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    nc.border = _thin(NAVY)

    # ── Sheet 2: 40-Year Projection ───────────────────────────────────────────
    ws_proj.cell(row=1, column=1, value="40-Year Annual Audit Ledger").font = F_TITLE
    ws_proj.row_dimensions[3].height = 30

    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        c = ws_proj.cell(row=3, column=ci, value=h)
        c.font = F_HDR; c.fill = _fill(NAVY); c.alignment = A_CTR; c.border = _thin()

    for ri, (_, row_data) in enumerate(df.iterrows(), 4):
        ws_proj.row_dimensions[ri].height = 18
        zebra = (ri % 2 == 0)
        for ci, val in enumerate(row_data, 1):
            cell = ws_proj.cell(row=ri, column=ci, value=val)
            cell.font = F_REG; cell.border = _thin()
            if zebra:
                cell.fill = _fill(LIGHT)
            col_name = headers[ci - 1]
            if col_name in ("Policy Year", "Age"):
                cell.alignment = A_CTR; cell.number_format = "0"
            elif col_name == "Sustainability Flag":
                cell.alignment = A_CTR
                is_exhausted = str(val) == "Corpus Exhausted"
                cell.fill = _fill(R_FILL if is_exhausted else G_FILL)
                cell.font = Font(name="Calibri", size=10, bold=True,
                                 color=R_TEXT if is_exhausted else G_TEXT)
            elif isinstance(val, (int, float, np.integer, np.floating)):
                # Display as Indian formatted string
                cell.value = indian_fmt(val)
                cell.alignment = A_RIGHT

    for col in ws_proj.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value and c.row > 2), default=10)
        ws_proj.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # ── Sheet 3: Charts ───────────────────────────────────────────────────────
    ws_gr.cell(row=1, column=1, value="VISUAL STRATEGIC ANCHORS").font = F_TITLE
    if os.path.exists(chart_path):
        img = openpyxl.drawing.image.Image(chart_path)
        img.anchor = "A3"
        ws_gr.add_image(img)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
# 4. PDF EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def build_pdf(df, summary, inputs, chart_path) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.6*inch, rightMargin=0.6*inch,
                            topMargin=0.6*inch, bottomMargin=0.6*inch)
    styles = getSampleStyleSheet()

    NAVY  = colors.HexColor("#1B365D")
    LIGHT = colors.HexColor("#F0F4F8")
    RED   = colors.HexColor("#9C0006")
    GREEN = colors.HexColor("#006100")

    title_style   = ParagraphStyle("Title2", parent=styles["Title"],
                                   textColor=NAVY, fontSize=18, spaceAfter=4)
    sub_style     = ParagraphStyle("Sub", parent=styles["Normal"],
                                   textColor=colors.HexColor("#555555"),
                                   fontSize=9, spaceAfter=12, leading=13)
    section_style = ParagraphStyle("Sec", parent=styles["Heading2"],
                                   textColor=NAVY, fontSize=11, spaceBefore=12)
    note_style    = ParagraphStyle("Note", parent=styles["Normal"],
                                   textColor=NAVY, fontSize=8.5, leading=13,
                                   backColor=colors.HexColor("#EFF4FB"), borderPad=6)

    elems = []

    elems += [
        Paragraph("Wealth Executive Summary", title_style),
        Paragraph(f"Strategic Advisory Report — {inputs['client_name']}", sub_style),
        HRFlowable(width="100%", thickness=1.5, color=NAVY),
        Spacer(1, 0.15*inch),
    ]

    if os.path.exists(chart_path):
        elems += [RLImage(chart_path, width=6.5*inch, height=3.0*inch), Spacer(1, 0.15*inch)]

    elems.append(Paragraph("Plan Performance Metrics", section_style))
    is_bad = summary["Sustainability Status"] != "Sustainable"

    kpi_data = [
        ["Metric", "Value"],
        ["Total Life Cover",              indian_fmt(summary["Life Cover Amount"])],
        ["Monthly Base SIP",              indian_fmt(summary["Monthly Insurance Payout"])],
        [f"Corpus at Age {inputs['swp_start_age']}", indian_fmt(summary["Corpus at SWP Start"])],
        ["Target Corpus for SWP",         indian_fmt(summary["Target Required Corpus"])],
        ["Surplus / Shortfall",           indian_fmt(summary["Surplus / Shortfall"])],
        ["Max Safe Monthly Withdrawal",   indian_fmt(summary["Max Safe Monthly SWP"])],
        ["Terminal Portfolio",            indian_fmt(summary["Final Year Corpus"])],
        ["Strategy Status",               summary["Sustainability Status"].upper()],
    ]

    status_color = colors.HexColor("#FFC7CE") if is_bad else colors.HexColor("#C6EFCE")
    kpi_table = Table(kpi_data, colWidths=[3.0*inch, 3.0*inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1,  0),  NAVY),
        ("TEXTCOLOR",      (0, 0), (-1,  0),  colors.white),
        ("FONTNAME",       (0, 0), (-1,  0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, -1),  9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),  [LIGHT, colors.white]),
        ("BACKGROUND",     (0, -1), (-1, -1), status_color),
        ("FONTNAME",       (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN",          (1, 0), (1, -1),   "RIGHT"),
        ("GRID",           (0, 0), (-1, -1),  0.4, colors.HexColor("#CCCCCC")),
        ("ROWHEIGHT",      (0, 0), (-1, -1),  18),
        ("TOPPADDING",     (0, 0), (-1, -1),  4),
        ("BOTTOMPADDING",  (0, 0), (-1, -1),  4),
    ]))
    elems += [kpi_table, Spacer(1, 0.2*inch)]

    elems += [
        Paragraph("💡 Advisory Note", section_style),
        Paragraph(summary["Client Presentation Note"], note_style),
        Spacer(1, 0.2*inch),
    ]

    elems.append(Paragraph("40-Year Chronological Audit Ledger", section_style))
    SHORT_COLS = [
        "Policy Year", "Age", "Total Monthly SIP (Rs.)",
        "Net End-of-Year Corpus (Rs.)", "Annual SWP Withdrawal (Rs.)", "Sustainability Flag"
    ]
    tbl_data = [SHORT_COLS]
    for _, row in df.iterrows():
        tbl_data.append([
            int(row["Policy Year"]),
            int(row["Age"]),
            indian_fmt(row["Total Monthly SIP (Rs.)"]),
            indian_fmt(row["Net End-of-Year Corpus (Rs.)"]),
            indian_fmt(row["Annual SWP Withdrawal (Rs.)"]),
            row["Sustainability Flag"],
        ])

    col_w = [0.7*inch, 0.5*inch, 1.3*inch, 1.5*inch, 1.3*inch, 1.2*inch]
    proj_table = Table(tbl_data, colWidths=col_w, repeatRows=1)
    proj_ts = TableStyle([
        ("BACKGROUND",    (0, 0), (-1,  0),  NAVY),
        ("TEXTCOLOR",     (0, 0), (-1,  0),  colors.white),
        ("FONTNAME",      (0, 0), (-1,  0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1),  7.5),
        ("ALIGN",         (0, 0), (-1, -1),  "CENTER"),
        ("ALIGN",         (2, 1), (4, -1),   "RIGHT"),
        ("GRID",          (0, 0), (-1, -1),  0.3, colors.HexColor("#CCCCCC")),
        ("ROWHEIGHT",     (0, 0), (-1, -1),  14),
        ("TOPPADDING",    (0, 0), (-1, -1),  2),
        ("BOTTOMPADDING", (0, 0), (-1, -1),  2),
    ])
    for i in range(1, len(tbl_data)):
        if i % 2 == 0:
            proj_ts.add("BACKGROUND", (0, i), (-1, i), LIGHT)
        status_val = tbl_data[i][-1]
        if status_val == "Corpus Exhausted":
            proj_ts.add("BACKGROUND", (5, i), (5, i), colors.HexColor("#FFC7CE"))
            proj_ts.add("TEXTCOLOR",  (5, i), (5, i), RED)
        else:
            proj_ts.add("BACKGROUND", (5, i), (5, i), colors.HexColor("#C6EFCE"))
            proj_ts.add("TEXTCOLOR",  (5, i), (5, i), GREEN)

    proj_table.setStyle(proj_ts)
    elems.append(proj_table)

    doc.build(elems)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
# 5. STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Wealth Architecture Engine", layout="wide", page_icon="📊")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@400;500;600&display=swap');

    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
    }

    [data-testid="stAppViewContainer"] {
        background: #0B1622;
    }
    [data-testid="stSidebar"] {
        background: #07101A;
        border-right: 1px solid #1C2D3E;
    }
    [data-testid="stSidebar"] * { color: #8BA3BF !important; }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] strong { color: #C5D8EE !important; }

    /* Page title */
    .dashboard-title {
        font-family: 'DM Serif Display', serif;
        font-size: 2rem;
        color: #E8F0FA;
        letter-spacing: -0.5px;
        line-height: 1.2;
        margin-bottom: 2px;
    }
    .dashboard-sub {
        font-size: 13px;
        color: #4A6A8A;
        letter-spacing: 0.8px;
        text-transform: uppercase;
        margin-bottom: 24px;
    }

    /* KPI Cards */
    .kpi-grid { display: flex; gap: 12px; margin-bottom: 24px; }
    .kpi-card {
        flex: 1;
        background: #111E2B;
        border: 1px solid #1C2D3E;
        border-top: 3px solid #4A90E2;
        border-radius: 8px;
        padding: 16px 18px;
        position: relative;
        overflow: hidden;
    }
    .kpi-card::after {
        content: '';
        position: absolute;
        top: 0; right: 0;
        width: 60px; height: 60px;
        background: radial-gradient(circle at top right, rgba(74,144,226,0.08), transparent);
        border-radius: 0 8px 0 0;
    }
    .kpi-card.accent { border-top-color: #E09B2A; }
    .kpi-card.accent::after { background: radial-gradient(circle at top right, rgba(224,155,42,0.08), transparent); }
    .kpi-card.good   { border-top-color: #27AE60; }
    .kpi-card.good::after { background: radial-gradient(circle at top right, rgba(39,174,96,0.08), transparent); }
    .kpi-card.bad    { border-top-color: #E74C3C; }
    .kpi-card.bad::after { background: radial-gradient(circle at top right, rgba(231,76,60,0.08), transparent); }

    .kpi-label {
        font-size: 10px;
        color: #4A6A8A;
        text-transform: uppercase;
        letter-spacing: 1px;
        font-weight: 600;
        margin-bottom: 6px;
    }
    .kpi-value {
        font-family: 'DM Serif Display', serif;
        font-size: 1.5rem;
        color: #E8F0FA;
        line-height: 1.1;
    }
    .kpi-sub {
        font-size: 11px;
        color: #3A5A7A;
        margin-top: 4px;
    }

    /* Section header */
    .section-header {
        font-size: 10px;
        font-weight: 700;
        color: #4A6A8A;
        text-transform: uppercase;
        letter-spacing: 2px;
        border-bottom: 1px solid #1C2D3E;
        padding-bottom: 8px;
        margin-bottom: 16px;
    }

    /* Note box */
    .note-box {
        background: #0F1E2D;
        border: 1px solid #1C2D3E;
        border-left: 4px solid #4A90E2;
        border-radius: 6px;
        padding: 16px 18px;
        color: #8BA3BF;
        font-size: 13px;
        line-height: 1.7;
    }

    /* Status badge */
    .badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 11px;
        font-weight: 700;
        letter-spacing: 0.5px;
        text-transform: uppercase;
    }
    .badge-green { background: rgba(39,174,96,0.15); color: #27AE60; border: 1px solid rgba(39,174,96,0.3); }
    .badge-red   { background: rgba(231,76,60,0.15);  color: #E74C3C; border: 1px solid rgba(231,76,60,0.3); }

    /* Download buttons */
    div[data-testid="stDownloadButton"] button {
        background: #111E2B !important;
        color: #8BA3BF !important;
        border: 1px solid #1C2D3E !important;
        border-radius: 6px !important;
        width: 100%;
        font-weight: 600;
        font-size: 13px;
        transition: all 0.2s;
    }
    div[data-testid="stDownloadButton"] button:hover {
        background: #1B365D !important;
        color: #E8F0FA !important;
        border-color: #4A90E2 !important;
    }

    /* Divider */
    hr { border-color: #1C2D3E !important; margin: 24px 0; }

    /* Dataframe */
    .stDataFrame { border: 1px solid #1C2D3E; border-radius: 8px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

# ── Page Header ───────────────────────────────────────────────────────────────
st.markdown('<div class="dashboard-title">📊 Wealth Architecture Engine</div>', unsafe_allow_html=True)
st.markdown('<div class="dashboard-sub">SIP · SWP · Step-Up · Life Cover · Indian Metric Projections</div>', unsafe_allow_html=True)

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 👤 Client Profile")
    client_name = st.text_input("Client Name", "Aditya Sharma")
    current_age = st.number_input("Current Age", 20, 75, 40)
    retire_age  = st.number_input("SWP Start Age", 40, 90, 60)

    st.markdown("---")
    st.markdown("### 📋 Policy Parameters")
    annual_prem = st.number_input("Annual Premium (Rs.)", 10_000, 10_000_000, 150_000, step=5_000)
    payout_pct  = st.slider("Insurance Payout %", 10, 100, 40) / 100
    life_mult   = st.slider("Life Cover Multiple", 1, 20, 7)
    ppt         = st.number_input("Premium Payment Term (yrs)", 1, 40, 12)
    policy_term = st.number_input("Policy Term (yrs)", 5, 50, 40)

    st.markdown("---")
    st.markdown("### 📈 Return & Withdrawal")
    exp_return  = st.slider("Expected Return (%)", 1.0, 25.0, 18.0, 0.5) / 100
    monthly_swp = st.number_input("Monthly SWP Target (Rs.)", 0, 2_000_000, 125_000, step=5_000)

    st.markdown("---")
    st.markdown("### 🪜 Manual SIP Step-Up Schedule")
    st.caption("Add year-specific absolute monthly SIP increases (Rs.)")

    if "step_up_rows" not in st.session_state:
        st.session_state.step_up_rows = []

    col_a, col_b = st.columns(2)
    new_yr  = col_a.number_input("Policy Year", 1, int(policy_term), 3, key="su_yr")
    new_amt = col_b.number_input("Monthly Rs. Increase", 0, 500_000, 1_000, step=500, key="su_amt")

    if st.button("➕ Add Step-Up", use_container_width=True):
        st.session_state.step_up_rows.append((int(new_yr), float(new_amt)))
        st.session_state.step_up_rows = sorted(
            list({yr: amt for yr, amt in st.session_state.step_up_rows}.items())
        )

    if st.session_state.step_up_rows:
        st.markdown("**Active Schedule:**")
        remove_yr = None
        for yr, amt in st.session_state.step_up_rows:
            c1, c2 = st.columns([3, 1])
            c1.markdown(f"Year {yr}: **{indian_fmt(amt)}/mo**")
            if c2.button("✕", key=f"rm_{yr}"):
                remove_yr = yr
        if remove_yr is not None:
            st.session_state.step_up_rows = [
                (y, a) for y, a in st.session_state.step_up_rows if y != remove_yr
            ]
            st.rerun()

    if st.button("🗑 Clear All Step-Ups", use_container_width=True):
        st.session_state.step_up_rows = []
        st.rerun()

# ── ENGINE ────────────────────────────────────────────────────────────────────
step_up_schedule = {yr: amt for yr, amt in st.session_state.get("step_up_rows", [])}

inputs = {
    "client_name":        client_name,
    "current_age":        current_age,
    "annual_premium":     annual_prem,
    "payout_pct":         payout_pct,
    "life_cover_multiple": life_mult,
    "ppt":                ppt,
    "policy_term":        policy_term,
    "expected_return":    exp_return,
    "monthly_swp":        monthly_swp,
    "swp_start_age":      retire_age,
}

engine  = GranularWealthEngine(inputs, step_up_schedule)
df      = engine.run_projection()
summary = engine.generate_summary(df)
is_bad  = summary["Sustainability Status"] != "Sustainable"

# ── KPI CARDS ─────────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">Performance Dashboard</div>', unsafe_allow_html=True)

surplus = summary["Surplus / Shortfall"]
surplus_class = "good" if surplus >= 0 else "bad"
status_badge  = (
    '<span class="badge badge-green">✓ Sustainable</span>' if not is_bad
    else '<span class="badge badge-red">⚠ Exhausted</span>'
)

k1, k2, k3, k4, k5 = st.columns(5)

def kpi_card(col, label, value, sub="", cls=""):
    col.markdown(
        f'<div class="kpi-card {cls}">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value}</div>'
        f'<div class="kpi-sub">{sub}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

kpi_card(k1, "Life Cover",             indian_fmt(summary["Life Cover Amount"]),
         f"{life_mult}× annual premium")
kpi_card(k2, "Monthly Base SIP",       indian_fmt(summary["Monthly Insurance Payout"]),
         f"{payout_pct*100:.0f}% of premium")
kpi_card(k3, f"Corpus at Age {retire_age}", indian_fmt(summary["Corpus at SWP Start"]),
         "at SWP start", "accent")
kpi_card(k4, "Surplus / Shortfall",    indian_fmt(surplus),
         "vs target corpus", surplus_class)
kpi_card(k5, "Max Safe SWP / Month",   indian_fmt(summary["Max Safe Monthly SWP"]),
         status_badge, "good" if not is_bad else "bad")

# ── CHART + ADVISORY ──────────────────────────────────────────────────────────
st.markdown("")
lc, rc = st.columns([3, 1])

with lc:
    st.markdown('<div class="section-header">Portfolio Growth Trajectory</div>', unsafe_allow_html=True)
    fig = build_chart(df, retire_age)
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)

with rc:
    st.markdown('<div class="section-header">Advisory Note</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="note-box">{summary["Client Presentation Note"]}</div>',
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    if is_bad:
        st.error(
            f"Corpus exhausts at Age **{summary['Age Corpus Exhausted']}** "
            f"(Policy Year {summary['Years Corpus Survives']})"
        )
    else:
        st.success("Strategy is fully sustainable through the policy term.")

    st.markdown(
        f"**Terminal Corpus:** {indian_fmt(summary['Final Year Corpus'])}",
        unsafe_allow_html=False,
    )

# ── PROJECTION TABLE ──────────────────────────────────────────────────────────
st.markdown("")
with st.expander("📋 View Full Projection Ledger", expanded=False):
    display_cols = [
        "Policy Year", "Age", "Total Monthly SIP (Rs.)",
        "New Step-Up Added (Rs.)", "Net End-of-Year Corpus (Rs.)",
        "Annual SWP Withdrawal (Rs.)", "Sustainability Flag",
    ]

    display_df = df[display_cols].copy()
    # Format numeric cols as Indian strings for display
    for col in ["Total Monthly SIP (Rs.)", "New Step-Up Added (Rs.)",
                "Net End-of-Year Corpus (Rs.)", "Annual SWP Withdrawal (Rs.)"]:
        display_df[col] = display_df[col].apply(indian_fmt)

    def color_status(val):
        if val == "Corpus Exhausted":
            return "background-color:#FFC7CE; color:#9C0006; font-weight:700"
        return "background-color:#C6EFCE; color:#006100; font-weight:700"

    styled = display_df.style.map(color_status, subset=["Sustainability Flag"])
    st.dataframe(styled, use_container_width=True, height=420)

# ── DOCUMENT GENERATION ───────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<div class="section-header">Document Generation</div>', unsafe_allow_html=True)

with tempfile.TemporaryDirectory() as tmp:
    chart_path = os.path.join(tmp, "chart.png")
    fig2 = build_chart(df, retire_age)
    fig2.savefig(chart_path, bbox_inches="tight", dpi=150, facecolor="#0F1923")
    plt.close(fig2)

    d1, d2 = st.columns(2)
    safe_name = client_name.replace(" ", "_")

    excel_bytes = build_excel(df, summary, inputs, step_up_schedule, chart_path)
    d1.download_button(
        label="📥 Download Styled Excel Workbook",
        data=excel_bytes,
        file_name=f"{safe_name}_Wealth_Plan.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    pdf_bytes = build_pdf(df, summary, inputs, chart_path)
    d2.download_button(
        label="📥 Download Executive PDF Report",
        data=pdf_bytes,
        file_name=f"{safe_name}_Wealth_Report.pdf",
        mime="application/pdf",
    )